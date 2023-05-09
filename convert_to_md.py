import openpyxl

# Load the workbook and select the first worksheet
xlsx_file = 'survey.xlsx'
wb = openpyxl.load_workbook(xlsx_file)
ws = wb.active

# Get the question names from the first row
question_names = [cell.value for cell in ws[1]]

# Open the Markdown file to write the contents
with open('output.md', 'w', encoding='utf-8') as md_file:
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Skip the first row (question names)
        if row_index == 1:
            continue

        # Write the respondent label to the Markdown file in bold
        respondent_label = f"**Respondent {row_index - 1}**"
        md_file.write(respondent_label + '\n\n')  # Added extra newline character

        responses = []
        for col_index, cell in enumerate(row):
            # Check if the cell is not empty
            if cell is not None:
                # Write the question name and the response to the Markdown file
                question = question_names[col_index]
                response = str(cell)
                responses.append(f"{question}:\n{response}\n")  # Added extra newline character

        # Write the responses to the Markdown file, separated by a newline
        md_file.write('\n'.join(responses) + '\n\n')
